import json
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter


STATUS_SYMBOLS = {
    "PASS": "○",
    "FAIL": "×",
    "MANUAL": "△",
    "NOT_IMPLEMENTED": "－",
    "ERROR": "!",
    "UNKNOWN": "?",
}


STATUS_COLORS = {
    "PASS": "C6EFCE",
    "FAIL": "FFC7CE",
    "MANUAL": "FFEB9C",
    "NOT_IMPLEMENTED": "D9EAF7",
    "ERROR": "F4CCCC",
    "UNKNOWN": "E7E6E6",
}


HEADER_FILL = PatternFill(
    fill_type="solid",
    fgColor="D9EAF7"
)


TITLE_FILL = PatternFill(
    fill_type="solid",
    fgColor="1F4E78"
)


THIN_BORDER = Border(
    left=Side(
        style="thin",
        color="BFBFBF"
    ),
    right=Side(
        style="thin",
        color="BFBFBF"
    ),
    top=Side(
        style="thin",
        color="BFBFBF"
    ),
    bottom=Side(
        style="thin",
        color="BFBFBF"
    ),
)


def normalize_server_data(
    server_data: Any
) -> dict:
    """
    Convert a server result object or dictionary
    into a standard dictionary.

    Supports:
        - dictionaries
        - objects with to_dict()
        - Windows server objects
        - Linux server dictionaries
    """

    if isinstance(server_data, dict):

        normalized = dict(server_data)

        normalized.setdefault(
            "hostname",
            "Unknown"
        )

        normalized.setdefault(
            "ip_addresses",
            []
        )

        normalized.setdefault(
            "execution_time",
            ""
        )

        normalized.setdefault(
            "operating_system",
            "Unknown"
        )

        normalized.setdefault(
            "requirements",
            {}
        )

        return normalized

    if hasattr(server_data, "to_dict"):

        normalized = server_data.to_dict()

        normalized.setdefault(
            "operating_system",
            getattr(
                server_data,
                "operating_system",
                "Windows"
            )
        )

        return normalized

    return {
        "hostname": getattr(
            server_data,
            "hostname",
            "Unknown"
        ),
        "ip_addresses": getattr(
            server_data,
            "ip_addresses",
            []
        ),
        "execution_time": getattr(
            server_data,
            "execution_time",
            ""
        ),
        "operating_system": getattr(
            server_data,
            "operating_system",
            "Windows"
        ),
        "requirements": getattr(
            server_data,
            "requirements",
            {}
        ),
    }


def load_report_requirements(
    metadata_path: str
) -> dict:
    """
    Load Japanese report requirement titles
    and display order.
    """

    path = Path(metadata_path)

    if not path.exists():
        raise FileNotFoundError(
            "Report metadata file not found: "
            f"{metadata_path}"
        )

    with path.open(
        "r",
        encoding="utf-8"
    ) as file:
        return json.load(file)


def get_requirement_status(
    server: dict,
    requirement_id: str
) -> str:
    """
    Return the overall status for one requirement
    on one server.
    """

    requirements = server.get(
        "requirements",
        {}
    )

    result = requirements.get(
        requirement_id
    )

    if result is None:
        return "NOT_IMPLEMENTED"

    if isinstance(result, str):
        return result

    return result.get(
        "status",
        "UNKNOWN"
    )


def apply_status_format(
    cell,
    status: str
) -> None:
    """
    Apply status symbol, background color,
    font, alignment, and border.
    """

    normalized_status = str(
        status
    ).upper()

    cell.value = STATUS_SYMBOLS.get(
        normalized_status,
        "?"
    )

    cell.fill = PatternFill(
        fill_type="solid",
        fgColor=STATUS_COLORS.get(
            normalized_status,
            STATUS_COLORS["UNKNOWN"]
        )
    )

    cell.font = Font(
        bold=True,
        size=14
    )

    cell.alignment = Alignment(
        horizontal="center",
        vertical="center"
    )

    cell.border = THIN_BORDER


def format_value(
    value: Any
) -> str:
    """
    Convert parser values into readable Excel text.
    """

    if value is None:
        return "未取得"

    if isinstance(value, bool):
        return (
            "True"
            if value
            else "False"
        )

    if isinstance(
        value,
        (list, tuple, set)
    ):
        return ", ".join(
            str(item)
            for item in value
        )

    if isinstance(value, dict):
        return json.dumps(
            value,
            ensure_ascii=False
        )

    return str(value)


def format_operator(
    operator: str
) -> str:
    """
    Convert internal evaluator operator names
    into readable symbols.
    """

    operator_map = {
        "equals": "=",
        "not_equals": "≠",
        "greater_than": ">",
        "greater_than_or_equal": "≥",
        "less_than": "<",
        "less_than_or_equal": "≤",
        "contains": "含む",
        "in": "いずれか",
    }

    return operator_map.get(
        operator,
        operator
    )


def create_summary_sheet(
    workbook: Workbook,
    servers: list[dict],
    metadata: dict
) -> None:
    """
    Create the customer-facing matrix report.

    Rows:
        Security requirements

    Columns:
        Server hostnames
    """

    worksheet = workbook.active
    worksheet.title = "設定状況"

    server_count = len(servers)

    last_column = 2 + server_count

    last_column_letter = get_column_letter(
        last_column
    )

    if server_count > 0:

        worksheet.merge_cells(
            start_row=1,
            start_column=3,
            end_row=1,
            end_column=last_column
        )

        title_cell = worksheet.cell(
            row=1,
            column=3,
            value="設定状況"
        )

        title_cell.font = Font(
            bold=True,
            size=18
        )

        title_cell.alignment = Alignment(
            horizontal="center",
            vertical="center"
        )

        title_cell.fill = HEADER_FILL
        title_cell.border = THIN_BORDER

    headers = [
        "No",
        "要件",
    ]

    for server in servers:

        headers.append(
            server.get(
                "hostname",
                "Unknown"
            )
        )

    for column_number, header in enumerate(
        headers,
        start=1
    ):

        cell = worksheet.cell(
            row=2,
            column=column_number,
            value=header
        )

        cell.font = Font(
            bold=True,
            size=11
        )

        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER

        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True
        )

    sorted_requirements = sorted(
        metadata.items(),
        key=lambda item: item[1].get(
            "order",
            999
        )
    )

    current_row = 3

    for (
        requirement_id,
        requirement_metadata
    ) in sorted_requirements:

        requirement_title = (
            requirement_metadata.get(
                "title",
                ""
            )
        )

        requirement_order = (
            requirement_metadata.get(
                "order",
                current_row - 2
            )
        )

        number_cell = worksheet.cell(
            row=current_row,
            column=1,
            value=requirement_order
        )

        requirement_cell = worksheet.cell(
            row=current_row,
            column=2,
            value=(
                f"{requirement_id} "
                f"{requirement_title}"
            )
        )

        number_cell.border = THIN_BORDER
        number_cell.alignment = Alignment(
            horizontal="center",
            vertical="center"
        )

        requirement_cell.border = THIN_BORDER
        requirement_cell.alignment = Alignment(
            horizontal="left",
            vertical="center",
            wrap_text=True
        )

        for server_index, server in enumerate(
            servers,
            start=3
        ):

            status = get_requirement_status(
                server,
                requirement_id
            )

            result_cell = worksheet.cell(
                row=current_row,
                column=server_index
            )

            apply_status_format(
                result_cell,
                status
            )

        worksheet.row_dimensions[
            current_row
        ].height = 48

        current_row += 1

    worksheet.column_dimensions["A"].width = 7
    worksheet.column_dimensions["B"].width = 68

    for column_number in range(
        3,
        last_column + 1
    ):

        worksheet.column_dimensions[
            get_column_letter(
                column_number
            )
        ].width = 18

    worksheet.row_dimensions[1].height = 32
    worksheet.row_dimensions[2].height = 35

    worksheet.freeze_panes = "C3"

    worksheet.auto_filter.ref = (
        f"A2:{last_column_letter}"
        f"{current_row - 1}"
    )

    legend_row = current_row + 1

    legend_title_cell = worksheet.cell(
        row=legend_row,
        column=1,
        value="凡例"
    )

    legend_title_cell.font = Font(
        bold=True
    )

    legend_values = [
        (
            "○",
            "PASS",
            "適合"
        ),
        (
            "×",
            "FAIL",
            "不適合"
        ),
        (
            "△",
            "MANUAL",
            "手動確認"
        ),
        (
            "－",
            "NOT_IMPLEMENTED",
            "未実装・未取得"
        ),
    ]

    for offset, (
        symbol,
        status,
        description
    ) in enumerate(
        legend_values,
        start=1
    ):

        row_number = (
            legend_row + offset
        )

        symbol_cell = worksheet.cell(
            row=row_number,
            column=1,
            value=symbol
        )

        symbol_cell.fill = PatternFill(
            fill_type="solid",
            fgColor=STATUS_COLORS[
                status
            ]
        )

        symbol_cell.font = Font(
            bold=True,
            size=13
        )

        symbol_cell.alignment = Alignment(
            horizontal="center",
            vertical="center"
        )

        symbol_cell.border = THIN_BORDER

        description_cell = worksheet.cell(
            row=row_number,
            column=2,
            value=description
        )

        description_cell.border = THIN_BORDER
        description_cell.alignment = Alignment(
            vertical="center"
        )


def create_details_sheet(
    workbook: Workbook,
    servers: list[dict],
    metadata: dict
) -> None:
    """
    Create the engineering detail worksheet
    for all Windows and Linux servers.
    """

    worksheet = workbook.create_sheet(
        "詳細結果"
    )

    headers = [
        "ホスト名",
        "OS",
        "IPアドレス",
        "要件",
        "要件内容",
        "設定項目",
        "比較条件",
        "期待値",
        "実際値",
        "結果",
    ]

    for column_number, header in enumerate(
        headers,
        start=1
    ):

        cell = worksheet.cell(
            row=1,
            column=column_number,
            value=header
        )

        cell.fill = TITLE_FILL

        cell.font = Font(
            bold=True,
            color="FFFFFF"
        )

        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True
        )

        cell.border = THIN_BORDER

    current_row = 2

    sorted_requirements = sorted(
        metadata.items(),
        key=lambda item: item[1].get(
            "order",
            999
        )
    )

    for server in servers:

        hostname = server.get(
            "hostname",
            "Unknown"
        )

        operating_system = server.get(
            "operating_system",
            "Windows"
        )

        ip_addresses = format_value(
            server.get(
                "ip_addresses",
                []
            )
        )

        requirements = server.get(
            "requirements",
            {}
        )

        for (
            requirement_id,
            report_metadata
        ) in sorted_requirements:

            requirement_result = (
                requirements.get(
                    requirement_id
                )
            )

            if requirement_result is None:
                continue

            title = report_metadata.get(
                "title",
                ""
            )

            if isinstance(
                requirement_result,
                str
            ):

                requirement_status = (
                    requirement_result
                )

                details = []

            else:

                requirement_status = (
                    requirement_result.get(
                        "status",
                        "UNKNOWN"
                    )
                )

                details = (
                    requirement_result.get(
                        "details",
                        []
                    )
                )

            if not details:

                values = [
                    hostname,
                    operating_system,
                    ip_addresses,
                    requirement_id,
                    title,
                    "",
                    "",
                    "",
                    "",
                    requirement_status,
                ]

                for column_number, value in enumerate(
                    values,
                    start=1
                ):

                    cell = worksheet.cell(
                        row=current_row,
                        column=column_number,
                        value=value
                    )

                    cell.border = THIN_BORDER

                    cell.alignment = Alignment(
                        vertical="top",
                        wrap_text=True
                    )

                apply_status_format(
                    worksheet.cell(
                        row=current_row,
                        column=10
                    ),
                    requirement_status
                )

                current_row += 1

                continue

            for detail in details:

                detail_status = detail.get(
                    "result",
                    "UNKNOWN"
                )

                values = [
                    hostname,
                    operating_system,
                    ip_addresses,
                    requirement_id,
                    title,
                    detail.get(
                        "setting",
                        ""
                    ),
                    format_operator(
                        detail.get(
                            "operator",
                            "equals"
                        )
                    ),
                    format_value(
                        detail.get(
                            "expected"
                        )
                    ),
                    format_value(
                        detail.get(
                            "actual"
                        )
                    ),
                    detail_status,
                ]

                for column_number, value in enumerate(
                    values,
                    start=1
                ):

                    cell = worksheet.cell(
                        row=current_row,
                        column=column_number,
                        value=value
                    )

                    cell.border = THIN_BORDER

                    cell.alignment = Alignment(
                        vertical="top",
                        wrap_text=True
                    )

                apply_status_format(
                    worksheet.cell(
                        row=current_row,
                        column=10
                    ),
                    detail_status
                )

                current_row += 1

    worksheet.freeze_panes = "A2"

    if current_row > 2:

        worksheet.auto_filter.ref = (
            f"A1:J{current_row - 1}"
        )

    column_widths = {
        "A": 22,
        "B": 26,
        "C": 20,
        "D": 10,
        "E": 55,
        "F": 35,
        "G": 15,
        "H": 25,
        "I": 25,
        "J": 12,
    }

    for (
        column_letter,
        width
    ) in column_widths.items():

        worksheet.column_dimensions[
            column_letter
        ].width = width

    worksheet.row_dimensions[1].height = 35


def generate_excel_report(
    server_results: list[Any],
    output_path: str,
    metadata_path: str = (
        "config/report_requirements.json"
    )
) -> str:
    """
    Generate one consolidated Excel report
    for multiple Windows and Linux servers.

    Parameters
    ----------
    server_results:
        List of Windows or Linux server results.

    output_path:
        Destination Excel workbook path.

    metadata_path:
        Japanese report requirement metadata file.

    Returns
    -------
    str:
        Generated workbook path.
    """

    if not server_results:
        raise ValueError(
            "At least one server result "
            "is required"
        )

    normalized_servers = [
        normalize_server_data(
            server
        )
        for server in server_results
    ]

    metadata = load_report_requirements(
        metadata_path
    )

    output_file = Path(
        output_path
    )

    output_file.parent.mkdir(
        parents=True,
        exist_ok=True
    )

    workbook = Workbook()

    create_summary_sheet(
        workbook,
        normalized_servers,
        metadata
    )

    create_details_sheet(
        workbook,
        normalized_servers,
        metadata
    )

    workbook.save(
        output_file
    )

    return str(
        output_file
    )